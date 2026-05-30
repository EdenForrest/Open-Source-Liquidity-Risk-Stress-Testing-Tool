"""
Export service — builds structured reports from pipeline results.

Formats:
  Excel (.xlsx) — 6 sheets: Summary, Liquidity Ladder, Stress Scenarios, Positions, Waterfall, Annex IV
  PDF           — A4 paginated report with tables (reportlab)
  XML           — ESMA AIFMD II Annex IV XML (replaces proprietary namespace)
"""
from __future__ import annotations

import io
import xml.etree.ElementTree as ET
from datetime import date
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette (matches dark-mode UI theme loosely, but Excel is light)
# ---------------------------------------------------------------------------
_HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")   # deep navy
_SECTION_FILL  = PatternFill("solid", fgColor="2D5F8A")   # mid-blue
_ALT_FILL      = PatternFill("solid", fgColor="EEF3F8")   # very light blue
_RED_FILL      = PatternFill("solid", fgColor="FFDAD9")
_AMBER_FILL    = PatternFill("solid", fgColor="FFF3CD")
_GREEN_FILL    = PatternFill("solid", fgColor="D4EDDA")

_WHITE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
_NORMAL_FONT   = Font(name="Calibri", size=10)
_TITLE_FONT    = Font(bold=True, name="Calibri", size=14, color="1E3A5F")

_THIN_SIDE     = Side(style="thin", color="CCCCCC")
_THIN_BORDER   = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)

_PCT_FMT  = "0.0%"
_EUR_FMT  = '#,##0'
_BPS_FMT  = '0" bps"'
_NUM_FMT  = "0.00"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hdr(ws, row: int, col: int, value: str, width: int | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _WHITE_FONT
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width and ws.column_dimensions[get_column_letter(col)].width < width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _sec(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    cell.fill = _SECTION_FILL
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _cell(ws, row: int, col: int, value: Any, fmt: str | None = None,
          bold: bool = False, fill: PatternFill | None = None,
          align: str = "left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Calibri", size=10)
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _flag_fill(flag: bool | None, warn: bool | None = None) -> PatternFill | None:
    if flag:
        return _RED_FILL
    if warn:
        return _AMBER_FILL
    return _GREEN_FILL


def _pct(v: float | None) -> float | None:
    return v  # already a fraction, format string adds %


def _eur(v: float | None) -> float | None:
    if v is None:
        return None
    return round(v)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(wb: openpyxl.Workbook, r: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    m = r.get("liquidity_metrics", {})
    aifmd = r.get("aifmd2", {})
    fund   = r.get("fund_name", "—")
    rdate  = r.get("reporting_date", str(date.today()))
    nav    = r.get("total_nav_eur", 0)

    # Title row
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = f"Liquidity Risk Report — {fund}"
    title.font  = _TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Reporting Date: {rdate}   |   Total NAV: EUR {nav:,.0f}   |   Generated: {date.today()}"
    sub.font  = Font(italic=True, color="666666", name="Calibri", size=10)
    sub.alignment = Alignment(horizontal="left", vertical="center")

    row = 4

    # --- Liquidity Metrics section ---
    ws.merge_cells(f"A{row}:F{row}")
    _sec(ws, row, 1, "  LIQUIDITY METRICS")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    row += 1

    _hdr(ws, row, 1, "Metric", 28)
    _hdr(ws, row, 2, "Value", 18)
    _hdr(ws, row, 3, "Status", 14)
    row += 1

    breach = m.get("breach_flag", False)
    warn   = m.get("warning_flag", False)
    overall_fill = _flag_fill(breach, warn)
    _cell(ws, row, 1, "Overall Status", bold=True)
    _cell(ws, row, 2, "BREACH" if breach else "WARNING" if warn else "OK",
          bold=True, fill=overall_fill, align="center")
    _cell(ws, row, 3, m.get("regulatory_basis", ""), fill=None)
    row += 1

    liq_rows = [
        ("T+0 Liquid (%NAV)",          m.get("t0_pct"),          _PCT_FMT,  False, False),
        ("T+0+1 Liquid (%NAV)",        m.get("t0_t1_pct"),        _PCT_FMT,  False, False),
        ("Illiquid (>T+7) (%NAV)",     m.get("illiquid_pct"),     _PCT_FMT,  False, False),
        ("Weighted Avg Days to Liquid", m.get("weighted_avg_days"), _NUM_FMT, False, False),
        ("Largest Position (%NAV)",    m.get("largest_position_pct"), _PCT_FMT, False, False),
        ("Swing Pricing Active",       m.get("swing_pricing_active"), None,   False, False),
        ("Gate Activated",             m.get("gate_activated"),    None,      m.get("gate_activated"), False),
        ("Suspension Activated",       m.get("suspension_activated"), None,   m.get("suspension_activated"), False),
        ("Swing Adjustment (%)",       m.get("swing_adjustment"),  _PCT_FMT,  False, False),
    ]
    for i, (label, val, fmt, is_breach, is_warn) in enumerate(liq_rows):
        fill = _ALT_FILL if i % 2 == 0 else None
        if is_breach:
            fill = _RED_FILL
        elif is_warn:
            fill = _AMBER_FILL
        _cell(ws, row, 1, label, fill=fill)
        _cell(ws, row, 2, val, fmt=fmt, fill=fill, align="right")
        _cell(ws, row, 3, "", fill=fill)
        row += 1

    row += 1

    # --- Geo section ---
    geo_warn   = m.get("geo_warning_flag", False)
    geo_breach = m.get("geo_breach_flag", False)
    if m.get("geo_top_country") is not None:
        ws.merge_cells(f"A{row}:F{row}")
        _sec(ws, row, 1, "  GEOGRAPHICAL CONCENTRATION (AIFMD Annex IV)")
        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
            ws.cell(row=row, column=c).border = _THIN_BORDER
        row += 1

        geo_fill = _flag_fill(geo_breach, geo_warn)
        geo_rows = [
            ("Geo Status",           "BREACH" if geo_breach else "WARNING" if geo_warn else "OK"),
            ("Top Country",          f"{m.get('geo_top_country','—')} ({m.get('geo_top_country_pct', 0)*100:.1f}%)"),
            ("EU Exposure (%NAV)",   f"{m.get('eu_pct', 0)*100:.1f}%"),
            ("Non-EU Exposure (%NAV)", f"{m.get('non_eu_pct', 0)*100:.1f}%"),
        ]
        for i, (label, val) in enumerate(geo_rows):
            fill = geo_fill if i == 0 else (_ALT_FILL if i % 2 == 0 else None)
            _cell(ws, row, 1, label, fill=fill)
            _cell(ws, row, 2, val, fill=fill, bold=(i == 0), align="right" if i > 0 else "center")
            _cell(ws, row, 3, "", fill=fill)
            row += 1

        # Top countries table
        top = m.get("top_countries", {})
        if top:
            _hdr(ws, row, 1, "Country", 12)
            _hdr(ws, row, 2, "% NAV", 12)
            _hdr(ws, row, 3, "EU?", 10)
            row += 1
            from liquidity_risk_tool.config.settings import EU_COUNTRIES
            for i, (cc, pct) in enumerate(sorted(top.items(), key=lambda x: -x[1])):
                fill = _ALT_FILL if i % 2 == 0 else None
                _cell(ws, row, 1, cc, fill=fill)
                _cell(ws, row, 2, pct, fmt=_PCT_FMT, fill=fill, align="right")
                _cell(ws, row, 3, "Yes" if cc in EU_COUNTRIES else "No", fill=fill, align="center")
                row += 1

        row += 1

    # --- AIFMD II section ---
    ws.merge_cells(f"A{row}:F{row}")
    _sec(ws, row, 1, "  AIFMD II COMPLIANCE")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    row += 1

    lev_breach = aifmd.get("leverage_breach", False)
    aifmd_rows = [
        ("Gross Leverage",             aifmd.get("gross_leverage"), _NUM_FMT, lev_breach, False),
        ("Commitment Leverage",        aifmd.get("commitment_leverage"), _NUM_FMT, lev_breach, False),
        ("Leverage Cap",               aifmd.get("leverage_cap"), _NUM_FMT, False, False),
        ("Leverage Breach",            aifmd.get("leverage_breach"), None, lev_breach, False),
        ("LMT Count",                  aifmd.get("lmt_count"), None, False, not aifmd.get("lmt_compliant", True)),
        ("LMT Compliant",              aifmd.get("lmt_compliant"), None, False, False),
        ("Loan Origination AIF",       aifmd.get("is_loan_origination_aif"), None, False, False),
        ("Borrower Breaches",          aifmd.get("borrower_breaches") or "None", None, bool(aifmd.get("borrower_breaches")), False),
    ]
    for i, (label, val, fmt, is_breach, is_warn) in enumerate(aifmd_rows):
        fill = _ALT_FILL if i % 2 == 0 else None
        if is_breach:
            fill = _RED_FILL
        elif is_warn:
            fill = _AMBER_FILL
        _cell(ws, row, 1, label, fill=fill)
        _cell(ws, row, 2, val, fmt=fmt, fill=fill, align="right")
        _cell(ws, row, 3, "", fill=fill)
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


def _build_liquidity_ladder(wb: openpyxl.Workbook, r: dict) -> None:
    ws = wb.create_sheet("Liquidity Ladder")
    ws.sheet_view.showGridLines = False

    normal = r.get("liquidity_ladder", [])
    stress = r.get("stress_ladder", [])

    headers = ["Bucket", "Market Value (EUR)", "Realisable Value (EUR)", "% NAV", "Cumulative % NAV"]
    row = 1

    # Normal ladder
    ws.merge_cells(f"A{row}:E{row}")
    _sec(ws, row, 1, "  Normal Conditions")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    row += 1

    for ci, h in enumerate(headers, 1):
        _hdr(ws, row, ci, h, 22)
    row += 1

    for i, bkt in enumerate(normal):
        fill = _ALT_FILL if i % 2 == 0 else None
        _cell(ws, row, 1, bkt.get("bucket"), bold=True, fill=fill)
        _cell(ws, row, 2, _eur(bkt.get("market_value_eur")), fmt=_EUR_FMT, fill=fill, align="right")
        _cell(ws, row, 3, _eur(bkt.get("realisable_value_eur")), fmt=_EUR_FMT, fill=fill, align="right")
        _cell(ws, row, 4, bkt.get("nav_pct"), fmt=_PCT_FMT, fill=fill, align="right")
        _cell(ws, row, 5, bkt.get("cumulative_nav_pct"), fmt=_PCT_FMT, fill=fill, align="right")
        row += 1

    row += 1

    # Stress ladder
    ws.merge_cells(f"A{row}:E{row}")
    _sec(ws, row, 1, "  Stress Conditions (Severe Combined scenario)")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    row += 1

    for ci, h in enumerate(headers, 1):
        _hdr(ws, row, ci, h, 22)
    row += 1

    for i, bkt in enumerate(stress):
        fill = _ALT_FILL if i % 2 == 0 else None
        _cell(ws, row, 1, bkt.get("bucket"), bold=True, fill=fill)
        _cell(ws, row, 2, _eur(bkt.get("market_value_eur")), fmt=_EUR_FMT, fill=fill, align="right")
        _cell(ws, row, 3, _eur(bkt.get("realisable_value_eur")), fmt=_EUR_FMT, fill=fill, align="right")
        _cell(ws, row, 4, bkt.get("nav_pct"), fmt=_PCT_FMT, fill=fill, align="right")
        _cell(ws, row, 5, bkt.get("cumulative_nav_pct"), fmt=_PCT_FMT, fill=fill, align="right")
        row += 1

    _set_col_widths(ws, [14, 24, 26, 14, 20])


def _build_stress_scenarios(wb: openpyxl.Workbook, r: dict) -> None:
    ws = wb.create_sheet("Stress Scenarios")
    ws.sheet_view.showGridLines = False

    results  = r.get("stress_results", [])
    metadata = {m["name"]: m for m in r.get("scenario_metadata", [])}

    headers = [
        "Scenario", "Equity Shock", "Credit Spread (bps)", "Rate Shock (bps)",
        "Haircut Mult.", "Redemption Rate",
        "NAV Impact (%)", "Stressed NAV (EUR)",
        "T+0+1 Coverage (%)", "Shortfall (EUR)", "Worst Case?",
    ]

    row = 1
    for ci, h in enumerate(headers, 1):
        _hdr(ws, row, ci, h, 18)
    row += 1

    for i, sc in enumerate(results):
        name       = sc.get("scenario_name", "")
        meta       = metadata.get(name, {})
        is_worst   = meta.get("is_worst_case", False)
        nav_impact = sc.get("nav_impact_pct", 0) or 0
        shortfall  = sc.get("shortfall_eur", 0) or 0
        coverage   = sc.get("t0_t1_coverage", sc.get("coverage_ratio"))

        fill = _RED_FILL if is_worst else (_ALT_FILL if i % 2 == 0 else None)

        _cell(ws, row, 1,  name, bold=is_worst, fill=fill)
        _cell(ws, row, 2,  meta.get("equity_shock"), fmt=_PCT_FMT, fill=fill, align="right")
        _cell(ws, row, 3,  meta.get("credit_spread_shock_bps"), fill=fill, align="right")
        _cell(ws, row, 4,  meta.get("rate_shock_bps"), fill=fill, align="right")
        _cell(ws, row, 5,  meta.get("liquidity_haircut_multiplier"), fmt=_NUM_FMT, fill=fill, align="right")
        _cell(ws, row, 6,  meta.get("redemption_rate"), fmt=_PCT_FMT, fill=fill, align="right")
        _cell(ws, row, 7,  nav_impact / 100 if nav_impact else None, fmt=_PCT_FMT, fill=fill, align="right")
        _cell(ws, row, 8,  _eur(sc.get("stressed_nav_eur")), fmt=_EUR_FMT, fill=fill, align="right")
        _cell(ws, row, 9,  coverage, fmt=_PCT_FMT, fill=fill, align="right")
        _cell(ws, row, 10, _eur(shortfall) if shortfall else None, fmt=_EUR_FMT, fill=fill, align="right")
        _cell(ws, row, 11, "Yes" if is_worst else "No", fill=fill, align="center")
        row += 1

    _set_col_widths(ws, [26, 14, 20, 16, 14, 16, 16, 22, 20, 18, 12])


def _build_positions(wb: openpyxl.Workbook, r: dict) -> None:
    ws = wb.create_sheet("Positions")
    ws.sheet_view.showGridLines = False

    positions = r.get("position_buckets", [])
    if not positions:
        ws["A1"].value = "No position data available."
        return

    # Determine available columns
    sample = positions[0]
    preferred_cols = [
        ("isin",              "ISIN",              16),
        ("name",              "Name",              30),
        ("asset_class",       "Asset Class",       20),
        ("country",           "Country",           10),
        ("currency",          "Currency",           10),
        ("bucket",            "Bucket",            10),
        ("market_value_eur",  "MV EUR",            18),
        ("realisable_value",  "Realisable (EUR)",  20),
        ("haircut",           "Haircut",           12),
        ("nav_pct",           "% NAV",             12),
        ("adv_participation", "ADV Part.",         14),
        ("days_to_liquidate", "Days to Liq.",      14),
    ]
    cols = [(k, lbl, w) for k, lbl, w in preferred_cols if k in sample]

    row = 1
    for ci, (_, lbl, w) in enumerate(cols, 1):
        _hdr(ws, row, ci, lbl, w)
        ws.column_dimensions[get_column_letter(ci)].width = w
    row += 1

    # Group by bucket for visual separation
    from collections import defaultdict
    buckets: dict[str, list] = defaultdict(list)
    bucket_order = ["T+0", "T+1", "T+3", "T+7", ">T+7"]
    for pos in positions:
        b = pos.get("bucket", "Other")
        buckets[b].append(pos)

    i = 0
    for bkt in bucket_order:
        if bkt not in buckets:
            continue
        # Section header
        ws.merge_cells(f"A{row}:{get_column_letter(len(cols))}{row}")
        _sec(ws, row, 1, f"  {bkt}")
        for c in range(2, len(cols) + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
            ws.cell(row=row, column=c).border = _THIN_BORDER
        row += 1

        for pos in buckets[bkt]:
            fill = _ALT_FILL if i % 2 == 0 else None
            for ci, (key, _, _) in enumerate(cols, 1):
                val = pos.get(key)
                fmt = None
                align = "left"
                if key in ("market_value_eur", "realisable_value"):
                    fmt, align = _EUR_FMT, "right"
                    val = _eur(val)
                elif key in ("haircut", "nav_pct", "adv_participation"):
                    fmt, align = _PCT_FMT, "right"
                elif key in ("days_to_liquidate",):
                    align = "right"
                _cell(ws, row, ci, val, fmt=fmt, fill=fill, align=align)
            i += 1
            row += 1


def _build_waterfall(wb: openpyxl.Workbook, r: dict) -> None:
    ws = wb.create_sheet("Waterfall")
    ws.sheet_view.showGridLines = False

    meta = r.get("waterfall_meta", {})
    daily = r.get("waterfall_summary", [])

    # Meta block
    row = 1
    ws.merge_cells(f"A{row}:D{row}")
    _sec(ws, row, 1, "  Waterfall Summary (Worst-Case Scenario)")
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    row += 1

    meta_rows = [
        ("Redemption Target (EUR)", _eur(meta.get("target_eur")), _EUR_FMT),
        ("Total Proceeds (EUR)",    _eur(meta.get("total_proceeds_eur")), _EUR_FMT),
        ("Target Met",              meta.get("target_met"), None),
        ("Days to Target",          meta.get("days_to_target"), None),
        ("Residual Shortfall (EUR)", _eur(meta.get("residual_shortfall_eur")), _EUR_FMT),
        ("NAV Before (EUR)",        _eur(meta.get("nav_before")), _EUR_FMT),
        ("NAV After (EUR)",         _eur(meta.get("nav_after")), _EUR_FMT),
        ("NAV Impact (%)",          meta.get("nav_impact_pct"), _PCT_FMT),
    ]
    for i, (label, val, fmt) in enumerate(meta_rows):
        fill = _ALT_FILL if i % 2 == 0 else None
        is_breach = label == "Target Met" and val is False
        if is_breach:
            fill = _RED_FILL
        _cell(ws, row, 1, label, bold=True, fill=fill)
        nav_impact_val = val / 100 if (label == "NAV Impact (%)" and val is not None) else val
        _cell(ws, row, 2, nav_impact_val, fmt=fmt, fill=fill, align="right")
        _cell(ws, row, 3, "", fill=fill)
        _cell(ws, row, 4, "", fill=fill)
        row += 1

    row += 1

    if not daily:
        ws["A" + str(row)].value = "No waterfall detail available."
        _set_col_widths(ws, [28, 22, 22, 22])
        return

    # Daily schedule
    ws.merge_cells(f"A{row}:D{row}")
    _sec(ws, row, 1, "  Daily Liquidation Schedule")
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    row += 1

    sample = daily[0]
    day_cols = [
        ("day",              "Day",               8),
        ("proceeds_eur",     "Proceeds (EUR)",    20),
        ("cumulative_eur",   "Cumulative (EUR)",  22),
        ("target_met",       "Target Met",        14),
        ("remaining_eur",    "Remaining (EUR)",   20),
    ]
    day_cols = [(k, lbl, w) for k, lbl, w in day_cols if k in sample]

    for ci, (_, lbl, w) in enumerate(day_cols, 1):
        _hdr(ws, row, ci, lbl, w)
        ws.column_dimensions[get_column_letter(ci)].width = w
    row += 1

    for i, d in enumerate(daily):
        fill = _GREEN_FILL if d.get("target_met") else (_ALT_FILL if i % 2 == 0 else None)
        for ci, (key, _, _) in enumerate(day_cols, 1):
            val = d.get(key)
            fmt = None
            align = "left"
            if key in ("proceeds_eur", "cumulative_eur", "remaining_eur"):
                fmt, align = _EUR_FMT, "right"
                val = _eur(val)
            elif key == "day":
                align = "center"
            _cell(ws, row, ci, val, fmt=fmt, fill=fill, align=align)
        row += 1

    _set_col_widths(ws, [28, 22, 22, 22, 22])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_excel(portfolio_results: dict, period: str | None = None) -> bytes:
    """Build and return an Excel workbook as bytes from a single-portfolio results dict."""
    wb = openpyxl.Workbook()

    _build_summary(wb, portfolio_results)
    _build_liquidity_ladder(wb, portfolio_results)
    _build_stress_scenarios(wb, portfolio_results)
    _build_positions(wb, portfolio_results)
    _build_waterfall(wb, portfolio_results)
    _annex_iv_sheet(wb, portfolio_results, period=period)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# PDF export (reportlab)
# ---------------------------------------------------------------------------

def build_pdf(r: dict) -> bytes:
    """Build and return a PDF report as bytes from a single-portfolio results dict."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    NAV_BLUE  = HexColor("#1E3A5F")
    MID_BLUE  = HexColor("#2D5F8A")
    ALT_GREY  = HexColor("#EEF3F8")
    RED_BG    = HexColor("#FFDAD9")
    AMBER_BG  = HexColor("#FFF3CD")
    GREEN_BG  = HexColor("#D4EDDA")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("rpt_title", parent=styles["Title"],
                                 textColor=NAV_BLUE, fontSize=16, spaceAfter=4)
    sub_style   = ParagraphStyle("rpt_sub",   parent=styles["Normal"],
                                 textColor=HexColor("#666666"), fontSize=9, spaceAfter=8)
    h2_style    = ParagraphStyle("rpt_h2",    parent=styles["Heading2"],
                                 textColor=NAV_BLUE, fontSize=11, spaceBefore=12, spaceAfter=4)

    m      = r.get("liquidity_metrics", {})
    aifmd  = r.get("aifmd2", {})
    fund   = r.get("fund_name", "—")
    rdate  = r.get("reporting_date", str(date.today()))
    nav    = r.get("total_nav_eur", 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=f"Liquidity Risk Report — {fund}",
        author="Liquidity Risk Tool",
    )

    COL_W = doc.width / 3

    def _tbl_style(header_rows=1, alt=True):
        cmds = [
            ("BACKGROUND", (0, 0), (-1, header_rows - 1), NAV_BLUE),
            ("TEXTCOLOR",  (0, 0), (-1, header_rows - 1), white),
            ("FONTNAME",   (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("FONTNAME",   (0, header_rows), (-1, -1), "Helvetica"),
            ("ROWBACKGROUND", (0, header_rows), (-1, -1), [white, ALT_GREY]),
            ("GRID",       (0, 0), (-1, -1), 0.25, HexColor("#CCCCCC")),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        return TableStyle(cmds)

    def _pf(v):
        if v is None:
            return "—"
        return f"{v * 100:.1f}%"

    def _ef(v):
        if v is None:
            return "—"
        return f"EUR {v:,.0f}"

    def _flag_color(breach, warn):
        if breach:
            return RED_BG
        if warn:
            return AMBER_BG
        return GREEN_BG

    story = []

    # --- Title ---
    story.append(Paragraph(f"Liquidity Risk Report — {fund}", title_style))
    story.append(Paragraph(
        f"Reporting Date: {rdate} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Total NAV: {_ef(nav)} &nbsp;&nbsp;|&nbsp;&nbsp; Generated: {date.today()}",
        sub_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=NAV_BLUE, spaceAfter=8))

    # --- Summary metrics ---
    breach = m.get("breach_flag", False)
    warn   = m.get("warning_flag", False)
    story.append(Paragraph("Liquidity Metrics", h2_style))

    summary_rows = [
        ["Metric", "Value", "Status"],
        ["Overall Status",
         "BREACH" if breach else "WARNING" if warn else "OK", ""],
        ["LCR T+1",           _pf(m.get("lcr_t1")),        ""],
        ["LCR T+3",           _pf(m.get("lcr_t3")),        ""],
        ["LCR T+7",           _pf(m.get("lcr_t7")),        ""],
        ["Illiquid (>T+7)",   _pf(m.get("illiquid_pct")),  ""],
        ["Days to 50% NAV",   f"{m.get('days_to_50pct', 0):.1f}" if m.get("days_to_50pct") else "—", ""],
        ["Days to 75% NAV",   f"{m.get('days_to_75pct', 0):.1f}" if m.get("days_to_75pct") else "—", ""],
        ["Days to 90% NAV",   f"{m.get('days_to_90pct', 0):.1f}" if m.get("days_to_90pct") else "—", ""],
        ["Top-10 Concentration", _pf(m.get("top10_concentration")), ""],
        ["Gate Activated",    str(m.get("gate_activated", False)), ""],
        ["Swing Pricing",     str(m.get("swing_pricing_active", False)), ""],
    ]
    tbl = Table(summary_rows, colWidths=[COL_W * 1.2, COL_W * 0.9, COL_W * 0.9])
    ts  = _tbl_style()
    ts.add("BACKGROUND", (0, 1), (-1, 1), _flag_color(breach, warn))
    if breach or warn:
        ts.add("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold")
    tbl.setStyle(ts)
    story.append(tbl)

    # --- Geo section ---
    if m.get("geo_top_country") is not None:
        geo_breach = m.get("geo_breach_flag", False)
        geo_warn   = m.get("geo_warning_flag", False)
        story.append(Paragraph("Geographical Concentration (AIFMD Annex IV)", h2_style))
        geo_rows = [
            ["Metric", "Value"],
            ["Geo Status",       "BREACH" if geo_breach else "WARNING" if geo_warn else "OK"],
            ["Top Country",      f"{m.get('geo_top_country','—')} ({m.get('geo_top_country_pct',0)*100:.1f}%)"],
            ["EU Exposure",      _pf(m.get("eu_pct"))],
            ["Non-EU Exposure",  _pf(m.get("non_eu_pct"))],
        ]
        top = m.get("top_countries", {})
        if top:
            geo_rows.append(["", ""])
            geo_rows.append(["Country", "% NAV"])
            try:
                from liquidity_risk_tool.config.settings import EU_COUNTRIES
            except Exception:
                EU_COUNTRIES = frozenset()
            for cc, pct in sorted(top.items(), key=lambda x: -x[1]):
                geo_rows.append([f"{cc} ({'EU' if cc in EU_COUNTRIES else 'Non-EU'})", _pf(pct)])
        tbl2 = Table(geo_rows, colWidths=[COL_W * 1.4, COL_W * 1.6])
        ts2  = _tbl_style()
        ts2.add("BACKGROUND", (0, 1), (-1, 1), _flag_color(geo_breach, geo_warn))
        tbl2.setStyle(ts2)
        story.append(tbl2)

    # --- Liquidity Ladder ---
    story.append(Paragraph("Liquidity Ladder", h2_style))
    normal = r.get("liquidity_ladder", [])
    stress = r.get("stress_ladder", [])

    def _ladder_rows(data, label):
        rows = [[label, "MV (EUR)", "Realisable (EUR)", "% NAV", "Cumul. % NAV"]]
        for bkt in data:
            rows.append([
                bkt.get("bucket", ""),
                _ef(bkt.get("market_value_eur")),
                _ef(bkt.get("realisable_value_eur")),
                _pf(bkt.get("nav_pct")),
                _pf(bkt.get("cumulative_nav_pct")),
            ])
        return rows

    if normal:
        ldr_w = doc.width / 5
        ldr = Table(_ladder_rows(normal, "Normal — Bucket"),
                    colWidths=[ldr_w * 0.7, ldr_w * 1.1, ldr_w * 1.2, ldr_w * 0.7, ldr_w * 1.3])
        ldr.setStyle(_tbl_style())
        story.append(ldr)
        story.append(Spacer(1, 6))

    if stress:
        ldr2 = Table(_ladder_rows(stress, "Stress — Bucket"),
                     colWidths=[ldr_w * 0.7, ldr_w * 1.1, ldr_w * 1.2, ldr_w * 0.7, ldr_w * 1.3])
        ts3  = _tbl_style()
        ts3.add("BACKGROUND", (0, 0), (-1, 0), MID_BLUE)
        ldr2.setStyle(ts3)
        story.append(ldr2)

    # --- Stress Scenarios ---
    story.append(Paragraph("Stress Scenarios", h2_style))
    sc_results  = r.get("stress_results", [])
    sc_meta_map = {m2["name"]: m2 for m2 in r.get("scenario_metadata", [])}
    if sc_results:
        sc_rows = [["Scenario", "NAV Impact", "T+1 Coverage", "Shortfall (EUR)", "Worst?"]]
        for sc in sc_results:
            name   = sc.get("scenario_name", "")
            meta2  = sc_meta_map.get(name, {})
            impact = sc.get("nav_impact_pct", 0) or 0
            sc_rows.append([
                name,
                f"{impact:.1f}%",
                _pf(sc.get("t0_t1_coverage") or sc.get("coverage_ratio")),
                _ef(sc.get("shortfall_eur") or 0),
                "Yes" if meta2.get("is_worst_case") else "No",
            ])
        sc_w = doc.width / 5
        sc_tbl = Table(sc_rows, colWidths=[sc_w * 1.4, sc_w * 0.8, sc_w * 0.9, sc_w * 1.1, sc_w * 0.8])
        ts4 = _tbl_style()
        for i, sc in enumerate(sc_results, 1):
            name = sc.get("scenario_name", "")
            if sc_meta_map.get(name, {}).get("is_worst_case"):
                ts4.add("BACKGROUND", (0, i), (-1, i), RED_BG)
        sc_tbl.setStyle(ts4)
        story.append(sc_tbl)

    # --- Positions ---
    story.append(Paragraph("Positions", h2_style))
    positions = r.get("position_buckets", [])
    if positions:
        pos_rows = [["ISIN", "Name", "Asset Class", "Bucket", "MV (EUR)", "Haircut", "Days"]]
        for pos in positions[:100]:  # cap at 100 rows for PDF readability
            pos_rows.append([
                pos.get("isin", ""),
                (pos.get("name") or "")[:28],
                pos.get("asset_class", ""),
                pos.get("bucket", ""),
                _ef(pos.get("market_value_eur")),
                _pf(pos.get("haircut")),
                f"{pos.get('days_to_liquidate', 0):.1f}" if pos.get("days_to_liquidate") is not None else "—",
            ])
        pos_w  = doc.width / 7
        pos_tbl = Table(pos_rows, colWidths=[pos_w * 0.9, pos_w * 1.5, pos_w * 1.0,
                                              pos_w * 0.6, pos_w * 1.1, pos_w * 0.7, pos_w * 0.7])
        pos_tbl.setStyle(_tbl_style())
        story.append(pos_tbl)
        if len(positions) > 100:
            story.append(Paragraph(
                f"<i>Note: {len(positions) - 100} additional positions omitted. Download Excel for the full list.</i>",
                ParagraphStyle("note", parent=styles["Normal"], fontSize=8, textColor=HexColor("#888888")),
            ))

    # --- AIFMD II ---
    story.append(Paragraph("AIFMD II — Directive (EU) 2024/927", h2_style))
    lev_breach = aifmd.get("leverage_breach", False)
    aifmd_rows = [
        ["Metric", "Value"],
        ["Gross Leverage",       f"{aifmd.get('gross_leverage', 0)*100:.1f}%" if aifmd.get("gross_leverage") is not None else "—"],
        ["Commitment Leverage",  f"{aifmd.get('commitment_leverage', 0)*100:.1f}%" if aifmd.get("commitment_leverage") is not None else "—"],
        ["Leverage Cap",         f"{aifmd.get('leverage_cap', 0)*100:.0f}%" if aifmd.get("leverage_cap") is not None else "—"],
        ["Leverage Breach",      str(lev_breach)],
        ["LMT Count",            str(aifmd.get("lmt_count", "—"))],
        ["LMT Compliant",        str(aifmd.get("lmt_compliant", False))],
    ]
    a_tbl = Table(aifmd_rows, colWidths=[COL_W * 1.4, COL_W * 1.6])
    ts5   = _tbl_style()
    if lev_breach:
        ts5.add("BACKGROUND", (0, 4), (-1, 4), RED_BG)
    a_tbl.setStyle(ts5)
    story.append(a_tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Annex IV Excel sheet
# ---------------------------------------------------------------------------

def _annex_iv_sheet(wb: openpyxl.Workbook, r: dict, period: str | None = None,
                    aifm_metadata: dict | None = None) -> None:
    """Append an 'Annex IV' sheet to wb containing ESMA AIFMD II Annex IV data."""
    from liquidity_risk_tool.reporting.annex_iv_mapper import build_annex_iv

    period_code = period or "2026Q1"
    data = build_annex_iv(r, period_code=period_code, aifm_metadata=aifm_metadata)

    ws = wb.create_sheet("Annex IV")

    def _title(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = _TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        return row + 1

    def _section(row, text):
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            c.fill = _SECTION_FILL
            if col == 1:
                c.value = text
                c.font = _WHITE_FONT
        return row + 1

    def _row(row, label, value, alt=False):
        fill = _ALT_FILL if alt else None
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _BOLD_FONT
        if fill:
            lc.fill = fill
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _NORMAL_FONT
        if fill:
            vc.fill = fill
        return row + 1

    row = 1
    row = _title(row, "ESMA AIFMD II — Annex IV Regulatory Reporting")
    row += 1

    # ---- Header info -------------------------------------------------------
    row = _section(row, "Reporting Period")
    row = _row(row, "Period Code", data["period_code"], alt=False)
    row = _row(row, "Period Type", data["period_type"], alt=True)
    row = _row(row, "Period Start", data["period_start"], alt=False)
    row = _row(row, "Period End", data["period_end"], alt=True)
    row = _row(row, "Schema Version", data["schema_version"], alt=False)
    row += 1

    # ---- AIFM identification -----------------------------------------------
    row = _section(row, "AIFM Identification")
    aifm = data["aifm"]
    row = _row(row, "AIFM Name", aifm["name"] or "(not set — set AIFM_NAME env var)", alt=False)
    row = _row(row, "AIFM LEI", aifm["lei"] or "(not set — set AIFM_LEI env var)", alt=True)
    row = _row(row, "AIFM National Code", aifm["national_code"] or "(not set)", alt=False)
    row = _row(row, "Reporting Member State", aifm["reporting_member_state"], alt=True)
    row += 1

    # ---- AIF identification ------------------------------------------------
    row = _section(row, "AIF Identification")
    aif = data["aif"]
    row = _row(row, "AIF Name", aif["name"], alt=False)
    row = _row(row, "AIF LEI", aif["lei"] or "(not set — set AIF_LEI env var)", alt=True)
    row = _row(row, "AIF National Code", aif["national_code"] or "(not set)", alt=False)
    row = _row(row, "Reporting Date", aif["reporting_date"], alt=True)
    row = _row(row, "Total NAV (EUR)", aif["total_nav_eur"], alt=False)
    ws.cell(row=row - 1, column=2).number_format = _EUR_FMT
    row += 1

    # ---- Asset liquidity profile -------------------------------------------
    row = _section(row, "Asset Liquidity Profile (ESMA Annex IV §43)")
    alp = data["asset_liquidity_profile"]
    horizon_labels = {
        "DAY_1":       "1 day (T+0/T+1)",
        "DAY_2_7":     "2–7 days (T+3/T+7)",
        "DAY_8_30":    "8–30 days",
        "DAY_31_90":   "31–90 days",
        "DAY_91_180":  "91–180 days",
        "DAY_181_365": "181–365 days",
        "DAY_MORE_365":"More than 365 days (>T+7)",
    }
    for i, (code, label) in enumerate(horizon_labels.items()):
        pct = alp.get(code, 0.0)
        row = _row(row, label, pct, alt=bool(i % 2))
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row += 1

    # ---- Investor liquidity profile ----------------------------------------
    row = _section(row, "Investor Liquidity Profile (ESMA Annex IV §44)")
    inv = data["investor_liquidity_profile"]
    row = _row(row, "Min Notice Period (days)", inv["min_notice_period_days"], alt=False)
    row = _row(row, "Redemption Frequency", inv["redemption_frequency"], alt=True)
    completeness = "Complete" if inv["data_complete"] else "Incomplete — share class data not available; conservative defaults applied"
    row = _row(row, "Data Completeness", completeness, alt=False)
    row += 1

    # ---- Geographical concentration ----------------------------------------
    row = _section(row, "Geographical Concentration (AIFMD Annex IV / ESMA34-39-897)")
    geo = data["geographical_concentration"]
    row = _row(row, "EU Exposure", geo["eu_pct"], alt=False)
    ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row = _row(row, "Non-EU Exposure", geo["non_eu_pct"], alt=True)
    ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row = _row(row, "Geo Status", geo["geo_status"], alt=False)
    row += 1

    # ---- Principal markets -------------------------------------------------
    row = _section(row, "Principal Markets Traded (Top 5)")
    _hdr(ws, row, 1, "Country Code")
    _hdr(ws, row, 2, "NAV %")
    row += 1
    for i, mkt in enumerate(data["principal_markets"]):
        alt = bool(i % 2)
        row = _row(row, mkt["country_code"], mkt["nav_pct"], alt=alt)
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row += 1

    # ---- Principal instruments ---------------------------------------------
    row = _section(row, "Principal Instruments Traded (Top 5)")
    for col, hdr_txt in enumerate(["ISIN", "Name", "Asset Class", "NAV %"], start=1):
        _hdr(ws, row, col, hdr_txt)
    row += 1
    for i, inst in enumerate(data["principal_instruments"]):
        alt = bool(i % 2)
        fill = _ALT_FILL if alt else None
        for col, val in enumerate([
            inst["isin"], inst["name"], inst["asset_class"], inst["nav_pct"]
        ], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _NORMAL_FONT
            if fill:
                c.fill = fill
        ws.cell(row=row, column=4).number_format = _PCT_FMT
        row += 1
    row += 1

    # ---- Portfolio concentrations ------------------------------------------
    row = _section(row, "Portfolio Concentrations by Asset Class")
    _hdr(ws, row, 1, "Asset Class")
    _hdr(ws, row, 2, "NAV %")
    row += 1
    for i, conc in enumerate(data["portfolio_concentrations"]):
        row = _row(row, conc["asset_class"], conc["nav_pct"], alt=bool(i % 2))
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row += 1

    # ---- Special arrangements / LMTs ---------------------------------------
    row = _section(row, "Special Arrangements / LMTs (AIFMD II Art. 16 & Annex V)")
    sa = data["special_arrangements"]
    row = _row(row, "Gate", "YES" if sa["gate"] else "No", alt=False)
    row = _row(row, "Suspension", "YES" if sa["suspension"] else "No", alt=True)
    row = _row(row, "Swing Pricing", "YES" if sa["swing_pricing"] else "No", alt=False)
    row = _row(row, "Side Pocket", "YES" if sa["side_pocket"] else "No", alt=True)
    row = _row(row, "LMT Count", sa["lmt_count"], alt=False)
    compliant_txt = "Compliant (≥2)" if sa["lmt_compliant"] else "NON-COMPLIANT (<2 required)"
    row = _row(row, "AIFMD II Art. 16 Compliance", compliant_txt, alt=True)
    lmt_cell = ws.cell(row=row - 1, column=2)
    lmt_cell.fill = _GREEN_FILL if sa["lmt_compliant"] else _RED_FILL
    row = _row(row, "Gate Threshold", sa["gate_threshold"], alt=False)
    ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row = _row(row, "Suspension Threshold", sa["suspension_threshold"], alt=True)
    ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row += 1

    # ---- Leverage ----------------------------------------------------------
    row = _section(row, "Leverage (AIFMD II Art. 15 / CDR 231/2013)")
    lev = data["leverage"]
    row = _row(row, "Gross Leverage", lev["gross_leverage"], alt=False)
    ws.cell(row=row - 1, column=2).number_format = _NUM_FMT
    row = _row(row, "Commitment Leverage", lev["commitment_leverage"], alt=True)
    ws.cell(row=row - 1, column=2).number_format = _NUM_FMT
    row = _row(row, "Leverage Cap", lev["leverage_cap"], alt=False)
    ws.cell(row=row - 1, column=2).number_format = _NUM_FMT
    breach_txt = "BREACH" if lev["leverage_breach"] else "Within cap"
    row = _row(row, "Cap Compliance", breach_txt, alt=True)
    ws.cell(row=row - 1, column=2).fill = _RED_FILL if lev["leverage_breach"] else _GREEN_FILL
    if lev.get("cap_utilization_pct") is not None:
        row = _row(row, "Cap Utilization", lev["cap_utilization_pct"], alt=False)
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row += 1

    # ---- Stress testing disclosure -----------------------------------------
    row = _section(row, "Stress Testing Disclosure (Annex IV §48)")
    sd = data["stress_disclosure"]
    row = _row(row, "Scenario", sd["scenario_name"], alt=False)
    row = _row(row, "NAV Before (EUR)", sd["nav_before_eur"], alt=True)
    if sd["nav_before_eur"] is not None:
        ws.cell(row=row - 1, column=2).number_format = _EUR_FMT
    row = _row(row, "NAV After (EUR)", sd["nav_after_eur"], alt=False)
    if sd["nav_after_eur"] is not None:
        ws.cell(row=row - 1, column=2).number_format = _EUR_FMT
    row = _row(row, "NAV Impact", sd["nav_impact_pct"], alt=True)
    if sd["nav_impact_pct"] is not None:
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row = _row(row, "Liquidity Before", sd["liquid_pct_before"], alt=False)
    if sd["liquid_pct_before"] is not None:
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row = _row(row, "Liquidity After", sd["liquid_pct_after"], alt=True)
    if sd["liquid_pct_after"] is not None:
        ws.cell(row=row - 1, column=2).number_format = _PCT_FMT
    row = _row(row, "Regulatory Basis", sd["regulatory_basis"], alt=False)
    row += 1

    # ---- Gaps / completeness -----------------------------------------------
    row = _section(row, "Completeness / Gap Analysis")
    gaps = data["gaps"]
    gap_labels = {
        "aifm_lei_missing":              "AIFM LEI",
        "aifm_national_code_missing":    "AIFM National Code",
        "aif_lei_missing":               "AIF LEI",
        "investor_profile_incomplete":   "Investor Liquidity Profile",
        "country_data_missing":          "Country Data",
        "leverage_missing":              "Leverage Computed",
        "stress_missing":                "Stress Scenarios",
    }
    for i, (key, label) in enumerate(gap_labels.items()):
        missing = gaps.get(key, False)
        status = "MISSING / INCOMPLETE" if missing else "OK"
        row = _row(row, label, status, alt=bool(i % 2))
        ws.cell(row=row - 1, column=2).fill = _RED_FILL if missing else _GREEN_FILL

    # ---- Column widths -----------------------------------------------------
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14


# ---------------------------------------------------------------------------
# Annex IV standalone Excel export (regulatory filing only — single sheet)
# ---------------------------------------------------------------------------

def build_excel_annex_iv(r: dict, period: str | None = None,
                         aifm_metadata: dict | None = None) -> bytes:
    """Return a single-sheet Excel workbook containing only the ESMA Annex IV
    regulatory data.  Distinct from build_excel() which is the full 6-sheet
    comprehensive internal risk management report."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    _annex_iv_sheet(wb, r, period=period, aifm_metadata=aifm_metadata)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# XML export — ESMA AIFMD II Annex IV
# ---------------------------------------------------------------------------

def build_xml(r: dict, period: str | None = None,
              aifm_metadata: dict | None = None) -> bytes:
    """Build and return an ESMA AIFMD II Annex IV XML report as bytes."""
    from liquidity_risk_tool.reporting.annex_iv_mapper import build_annex_iv

    period_code = period or "2026Q1"
    data = build_annex_iv(r, period_code=period_code, aifm_metadata=aifm_metadata)

    XMLNS = "urn:eu.europa.esma:aifmd:annex-iv:v1.2"

    def _el(parent, tag, text=None, **attrs):
        e = ET.SubElement(parent, tag, {k: str(v) for k, v in attrs.items() if v is not None})
        if text is not None:
            e.text = str(text) if text != "" else ""
        return e

    def _pct(v):
        if v is None:
            return ""
        try:
            return f"{float(v) * 100:.4f}"
        except (TypeError, ValueError):
            return ""

    def _num(v, d=4):
        if v is None:
            return ""
        try:
            return f"{float(v):.{d}f}"
        except (TypeError, ValueError):
            return str(v)

    period_type_map = {
        "Q1": "QUARTERLY", "Q2": "QUARTERLY", "Q3": "QUARTERLY", "Q4": "QUARTERLY",
        "H1": "SEMI_ANNUAL", "H2": "SEMI_ANNUAL",
        "Annual": "ANNUAL",
    }
    esma_period_type = period_type_map.get(data["period_type"], "QUARTERLY")

    root = ET.Element("AIFMD", {
        "xmlns": XMLNS,
        "version": data["schema_version"],
        "generatedDate": str(date.today()),
    })

    # ---- Header ------------------------------------------------------------
    hdr = ET.SubElement(root, "Header")
    _el(hdr, "CreationDateAndTime", str(date.today()))
    _el(hdr, "ConferringAuthorityCode", data["reporting_member_state"])
    _el(hdr, "FilingType", "INIT")
    _el(hdr, "AIFMNationalCode", data["aifm"]["national_code"] or "")
    _el(hdr, "AIFMLEICode", data["aifm"]["lei"] or "")
    _el(hdr, "ReportingPeriodStartDate", data["period_start"])
    _el(hdr, "ReportingPeriodEndDate", data["period_end"])
    _el(hdr, "ReportingPeriodType", esma_period_type)
    _el(hdr, "SchemaVersion", data["schema_version"])

    # ---- AIFRecordInfo -----------------------------------------------------
    aif_rec = ET.SubElement(root, "AIFRecordInfo")

    # AIF identification
    aif_id = ET.SubElement(aif_rec, "AIFIdentification")
    _el(aif_id, "AIFNationalCode", data["aif"]["national_code"] or "")
    _el(aif_id, "AIFLEICode", data["aif"]["lei"] or "")
    _el(aif_id, "AIFName", data["aif"]["name"])
    _el(aif_id, "ReportingDate", data["aif"]["reporting_date"])
    _el(aif_id, "TotalNAVEUR", _num(data["aif"]["total_nav_eur"], 2))

    # Principal markets
    pm_el = ET.SubElement(aif_rec, "PrincipalMarketsTraded")
    for mkt in data["principal_markets"]:
        m = ET.SubElement(pm_el, "Market")
        _el(m, "CountryCode", mkt["country_code"])
        _el(m, "NavPct", _pct(mkt["nav_pct"]))

    # Principal instruments
    pi_el = ET.SubElement(aif_rec, "PrincipalInstrumentsTraded")
    for inst in data["principal_instruments"]:
        i = ET.SubElement(pi_el, "Instrument")
        _el(i, "ISIN", inst["isin"])
        _el(i, "Name", inst["name"])
        _el(i, "AssetClass", inst["asset_class"])
        _el(i, "NavPct", _pct(inst["nav_pct"]))

    # Portfolio concentrations
    conc_el = ET.SubElement(aif_rec, "PortfolioConcentrations")
    for c in data["portfolio_concentrations"]:
        ce = ET.SubElement(conc_el, "Concentration")
        _el(ce, "AssetClass", c["asset_class"])
        _el(ce, "NavPct", _pct(c["nav_pct"]))

    # Asset liquidity profile (ESMA 7-bucket)
    alp_el = ET.SubElement(aif_rec, "AssetLiquidityProfile")
    alp = data["asset_liquidity_profile"]
    _el(alp_el, "Day1NavPct",        _pct(alp.get("DAY_1", 0.0)))
    _el(alp_el, "Day2To7NavPct",     _pct(alp.get("DAY_2_7", 0.0)))
    _el(alp_el, "Day8To30NavPct",    _pct(alp.get("DAY_8_30", 0.0)))
    _el(alp_el, "Day31To90NavPct",   _pct(alp.get("DAY_31_90", 0.0)))
    _el(alp_el, "Day91To180NavPct",  _pct(alp.get("DAY_91_180", 0.0)))
    _el(alp_el, "Day181To365NavPct", _pct(alp.get("DAY_181_365", 0.0)))
    _el(alp_el, "DayMore365NavPct",  _pct(alp.get("DAY_MORE_365", 0.0)))

    # Investor liquidity profile
    inv = data["investor_liquidity_profile"]
    inv_el = ET.SubElement(aif_rec, "InvestorLiquidityProfile")
    _el(inv_el, "MinNoticePeriodDays", str(inv["min_notice_period_days"]))
    _el(inv_el, "RedemptionFrequency", inv["redemption_frequency"])
    _el(inv_el, "DataComplete", "true" if inv["data_complete"] else "false")

    # Geographical concentration
    geo = data["geographical_concentration"]
    geo_el = ET.SubElement(aif_rec, "GeographicalConcentration")
    _el(geo_el, "EUNavPct",    _pct(geo["eu_pct"]))
    _el(geo_el, "NonEUNavPct", _pct(geo["non_eu_pct"]))
    _el(geo_el, "GeoStatus",   geo["geo_status"])
    for c in geo["top_countries"]:
        ce = ET.SubElement(geo_el, "Country")
        _el(ce, "Code", c["country_code"])
        _el(ce, "NavPct", _pct(c["nav_pct"]))

    # Special arrangements / LMTs
    sa = data["special_arrangements"]
    sa_el = ET.SubElement(aif_rec, "SpecialArrangements")
    _el(sa_el, "Gate",            "true" if sa["gate"] else "false")
    _el(sa_el, "Suspension",      "true" if sa["suspension"] else "false")
    _el(sa_el, "SwingPricing",    "true" if sa["swing_pricing"] else "false")
    _el(sa_el, "SidePocket",      "true" if sa["side_pocket"] else "false")
    _el(sa_el, "LMTCount",        str(sa["lmt_count"]))
    _el(sa_el, "LMTCompliant",    "true" if sa["lmt_compliant"] else "false")
    _el(sa_el, "GateThreshold",   _num(sa["gate_threshold"]))
    _el(sa_el, "SuspensionThreshold", _num(sa["suspension_threshold"]))
    lmts_el = ET.SubElement(sa_el, "LMTList")
    for lmt in sa["lmt_list"]:
        _el(lmts_el, "LMT", lmt)

    # Leverage (Art. 15 AIFMD II)
    lev = data["leverage"]
    lev_el = ET.SubElement(aif_rec, "LeverageInfo")
    _el(lev_el, "GrossLeverage",      _num(lev["gross_leverage"]))
    _el(lev_el, "CommitmentLeverage", _num(lev["commitment_leverage"]))
    _el(lev_el, "LeverageCap",        _num(lev["leverage_cap"]))
    _el(lev_el, "CapBreach",          "true" if lev["leverage_breach"] else "false")
    if lev.get("cap_utilization_pct") is not None:
        _el(lev_el, "CapUtilizationPct", _pct(lev["cap_utilization_pct"]))

    # Stress testing disclosure — ESMA Annex IV §48 requires only whether stress
    # tests are performed and the regulatory basis, not scenario parameters.
    sd = data["stress_disclosure"]
    st_el = ET.SubElement(aif_rec, "StressTestingDisclosure")
    _el(st_el, "StressTestsPerformed", "false" if not sd.get("scenario_name") else "true")
    _el(st_el, "RegulatoryBasis",      sd["regulatory_basis"])

    # Gaps / completeness
    gaps_el = ET.SubElement(aif_rec, "CompletenessFlags")
    for key, missing in data["gaps"].items():
        _el(gaps_el, key, "MISSING" if missing else "OK")

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    buf = io.BytesIO()
    tree.write(buf, encoding="utf-8", xml_declaration=True)
    buf.seek(0)
    return buf.read()
