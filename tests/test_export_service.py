"""
Characterization tests for backend.services.export_service.

Locks down the current output shape of the four public export functions
(build_excel, build_pdf, build_excel_annex_iv, build_xml) against the demo
pipeline dataset, so later refactor phases (Phase 2-5 of the consolidation
plan) can be verified as non-regressing. These tests do not diff exact
byte content (binary formats + embedded timestamps make that brittle) —
instead they assert structural invariants: sheet names/counts, non-empty
bytes, well-formed XML with the expected ESMA namespace and element tree.
"""
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from backend.services.export_service import (
    build_excel,
    build_excel_annex_iv,
    build_pdf,
    build_xml,
)
from backend.services.pipeline_service import run_full_pipeline

ROOT = Path(__file__).parent.parent
HOLDINGS_PATH = str(ROOT / "data" / "HOLDINGS_20260515001555.csv")
NAV_PATH = str(ROOT / "data" / "NAV_20260515001555.csv")
MARKET_DATA_PATH = str(ROOT / "data" / "market_data_ALL.csv")

XMLNS = "urn:eu.europa.esma:aifmd:annex-iv:v1.2"

FULL_WORKBOOK_SHEETS = [
    "Summary", "Liquidity Ladder", "Stress Scenarios",
    "Positions", "Waterfall", "Annex IV",
]


@pytest.fixture(scope="module")
def result():
    return run_full_pipeline(HOLDINGS_PATH, NAV_PATH, MARKET_DATA_PATH)


class TestBuildExcel:
    def test_returns_nonempty_bytes(self, result):
        out = build_excel(result)
        assert isinstance(out, bytes)
        assert len(out) > 0

    def test_has_exactly_six_sheets_in_order(self, result):
        out = build_excel(result)
        wb = openpyxl.load_workbook(BytesIO(out))
        assert wb.sheetnames == FULL_WORKBOOK_SHEETS

    def test_summary_sheet_has_title(self, result):
        out = build_excel(result)
        wb = openpyxl.load_workbook(BytesIO(out))
        ws = wb["Summary"]
        assert result["fund_name"] in ws["A1"].value

    def test_annex_iv_sheet_nonempty(self, result):
        out = build_excel(result)
        wb = openpyxl.load_workbook(BytesIO(out))
        ws = wb["Annex IV"]
        assert ws.max_row > 1


class TestBuildExcelAnnexIv:
    def test_returns_nonempty_bytes(self, result):
        out = build_excel_annex_iv(result)
        assert isinstance(out, bytes)
        assert len(out) > 0

    def test_single_sheet_named_annex_iv(self, result):
        out = build_excel_annex_iv(result)
        wb = openpyxl.load_workbook(BytesIO(out))
        assert wb.sheetnames == ["Annex IV"]

    def test_accepts_period_override(self, result):
        out = build_excel_annex_iv(result, period="2026Q2")
        wb = openpyxl.load_workbook(BytesIO(out))
        ws = wb["Annex IV"]
        values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
        assert "2026Q2" in values


class TestBuildPdf:
    def test_returns_nonempty_bytes(self, result):
        out = build_pdf(result)
        assert isinstance(out, bytes)
        assert len(out) > 0

    def test_starts_with_pdf_magic_bytes(self, result):
        out = build_pdf(result)
        assert out[:5] == b"%PDF-"


def _qn(tag):
    """Namespace-qualify a bare tag name for ElementTree find/findall calls.

    build_xml's root declares a default xmlns, so every element in the tree
    is parsed into that namespace (root.tag becomes "{ns}AIFMD", not the bare
    "AIFMD") and unqualified find("Header") paths silently return None.
    """
    return f"{{{XMLNS}}}{tag}"


class TestBuildXml:
    def test_returns_nonempty_bytes(self, result):
        out = build_xml(result)
        assert isinstance(out, bytes)
        assert len(out) > 0

    def test_well_formed_with_expected_namespace(self, result):
        out = build_xml(result)
        root = ET.fromstring(out)
        assert root.tag == _qn("AIFMD")
        assert root.get("version") is not None

    def test_has_header_and_aif_record_info_children(self, result):
        out = build_xml(result)
        root = ET.fromstring(out)
        child_tags = {c.tag for c in root}
        assert child_tags == {_qn("Header"), _qn("AIFRecordInfo")}

    def test_header_contains_period_dates(self, result):
        out = build_xml(result)
        root = ET.fromstring(out)
        header = root.find(_qn("Header"))
        assert header.find(_qn("ReportingPeriodStartDate")) is not None
        assert header.find(_qn("ReportingPeriodEndDate")) is not None
        assert header.find(_qn("ReportingPeriodType")).text in {
            "QUARTERLY", "SEMI_ANNUAL", "ANNUAL",
        }

    def test_aif_record_info_has_asset_liquidity_profile_seven_buckets(self, result):
        out = build_xml(result)
        root = ET.fromstring(out)
        alp = root.find(f"{_qn('AIFRecordInfo')}/{_qn('AssetLiquidityProfile')}")
        assert alp is not None
        expected_children = {
            "Day1NavPct", "Day2To7NavPct", "Day8To30NavPct", "Day31To90NavPct",
            "Day91To180NavPct", "Day181To365NavPct", "DayMore365NavPct",
        }
        assert {c.tag for c in alp} == {_qn(t) for t in expected_children}

    def test_special_arrangements_lmt_compliance_present(self, result):
        out = build_xml(result)
        root = ET.fromstring(out)
        sa = root.find(f"{_qn('AIFRecordInfo')}/{_qn('SpecialArrangements')}")
        assert sa is not None
        assert sa.find(_qn("LMTCount")) is not None
        assert sa.find(_qn("LMTCompliant")).text in {"true", "false"}

    def test_accepts_period_override(self, result):
        out = build_xml(result, period="2026Q3")
        root = ET.fromstring(out)
        # Q3 falls in the QUARTERLY bucket per export_service's period_type_map.
        header = root.find(_qn("Header"))
        assert header.find(_qn("ReportingPeriodType")).text == "QUARTERLY"
